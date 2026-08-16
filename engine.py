"""
engine.py — ליבת מערכת החיובים של ג.ד. פיקה הולדינגס בע"מ (ח.פ 515980514)

גלגול-ווב של build/sim-billing/sim_billing.py (my-aios, 06/08/2026):
קובץ חיובים (שם, טלפון, סכום) → ולידציה → חשבונית מס/קבלה בחשבונית ירוקה
לכל שורה → SMS/וואטסאפ עם לינק למסמך.

עקרונות שלא זזים:
- שער בטיחות: לפני כל הפקה מאמתים שהחשבון המחובר בחשבונית ירוקה הוא ג.ד. פיקה
  (EXPECTED_TAX_ID). חשבון אחר ⇒ עצירה, שום מסמך לא נוצר.
- מניעת חיוב כפול: מקור האמת הוא הספרים עצמם — לפני ריצה שולפים מחשבונית ירוקה
  את כל מסמכי החודש ומדלגים על טלפון שכבר קיבל מסמך. עמיד גם ב-redeploy בענן
  (אין תלות בדיסק מקומי).
- ההודעה ללקוח נשלחת רק אחרי שהמסמך הופק בהצלחה.

כל הסודות ב-env בלבד: SIM_BILLING_GI_KEY/SECRET, INFORU_USER/TOKEN,
GREEN_API_BOT_INSTANCE_ID/TOKEN.
"""

from __future__ import annotations
import os
import io
import re
import csv
import calendar
from datetime import datetime
from dataclasses import dataclass, field

import requests

try:
    from zoneinfo import ZoneInfo
    IL_TZ = ZoneInfo("Asia/Jerusalem")
except Exception:
    IL_TZ = None

GI_BASE = os.getenv("SIM_BILLING_GI_BASE", "https://api.greeninvoice.co.il/api/v1").strip()
EXPECTED_TAX_ID = os.getenv("EXPECTED_TAX_ID", "515980514").strip()
BUSINESS_NAME = "ג.ד. פיקה הולדינגס בע\"מ"

DOC_TYPE = int(os.getenv("DOC_TYPE", "320"))           # 320 = חשבונית מס/קבלה
PAYMENT_TYPE = int(os.getenv("PAYMENT_TYPE", "4"))     # 4 = הוראת קבע/העברה
VAT_RATE = float(os.getenv("VAT_RATE", "0.18"))
VAT_MODE = os.getenv("VAT_MODE", "included").strip()   # included | add | exempt
LINE_DESCRIPTION = os.getenv("LINE_DESCRIPTION", "דמי קו סלולר - חודש {month}")
DOC_REMARKS = os.getenv("DOC_REMARKS", "חיוב חודשי בהוראת קבע - שירותי סלולר")
SMS_SENDER = os.getenv("SMS_SENDER", "GD-PIKA")
MESSAGE_TEMPLATE = os.getenv("MESSAGE_TEMPLATE",
    "שלום {name}, חשבונך חויב ב-{amount} ₪ עבור קו הסלולר לחודש {month}. "
    "המסמך שלך: {link}\n"
    "Hello {name}, your account was charged {amount} ILS for your mobile line "
    "({month}). Your document: {link}\nG.D. Pika Holdings")

# ל-SMS: הודעה קומפקטית באנגלית (עברית כופה קידוד UCS-2 — 70 תווים למקטע במקום
# 160 — והתבנית המלאה הייתה עולה ~5 מקטעים בתשלום לכל עובד)
SMS_MESSAGE_TEMPLATE = os.getenv("SMS_MESSAGE_TEMPLATE",
    "G.D. Pika: {name}, your mobile line receipt for {month} ({amount} ILS): {link}")


def build_message(name: str, amount, month: str, link: str) -> str:
    tpl = SMS_MESSAGE_TEMPLATE if channel() == "inforu" else MESSAGE_TEMPLATE
    return tpl.format(name=name, amount=amount, month=month, link=link)

# ערוץ שליחה: dry (בלי הודעות) | inforu (SMS) | whatsapp (Green API)
def channel() -> str:
    return os.getenv("SEND_CHANNEL", "dry").strip().lower()


COLUMN_ALIASES = {
    "passport": ["מספר דרכון", "דרכון", "passport_number", "passport", "passport no",
                 "passport_no", "passport number"],
    "name":   ["שם מלא", "שם", "שם עובד", "שם העובד", "שם הלקוח", "לקוח", "עובד",
               "full_name", "full name", "name", "worker", "customer"],
    "phone":  ["מספר טלפון", "טלפון", "נייד", "מס טלפון", "מס' טלפון",
               "phone_number", "phone number", "phone", "mobile"],
    "amount": ["סכום", "מחיר", "סכום חיוב", "סכום לחיוב", "amount", "price"],
    "gmt":    ["מספר חשבון gmt", "חשבון gmt", "gmt_account_number", "gmt_account",
               "gmt account", "gmt", "מספר ארנק gmt", "ארנק gmt"],
}


# ---------------------------------------------------------------- קריאת קובץ

def read_rows_from_bytes(blob: bytes, filename: str):
    """CSV/XLSX → (header, rows). זורק ValueError עם הסבר בעברית."""
    if filename.lower().endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
        ws = wb.active
        rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
        wb.close()
    else:
        text = None
        for enc in ("utf-8-sig", "cp1255", "utf-8"):
            try:
                text = blob.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise ValueError("לא הצלחתי לקרוא את קידוד הקובץ — לשמור כ-CSV UTF-8 או Excel")
        rows = list(csv.reader(text.splitlines()))
    rows = [r for r in rows if any(str(c).strip() for c in r if c is not None)]
    if not rows:
        raise ValueError("הקובץ ריק")
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    return header, rows[1:]


def map_columns(header):
    mapping = {}
    lower = [h.strip().lower() for h in header]
    for fieldname, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias.lower() in lower:
                mapping[fieldname] = lower.index(alias.lower())
                break
    missing = [f for f in ("passport", "name", "phone", "amount", "gmt") if f not in mapping]
    if missing:
        heb = {"passport": "מספר דרכון", "name": "שם מלא", "phone": "מספר טלפון",
               "amount": "סכום", "gmt": "מספר חשבון GMT"}
        raise ValueError(
            f"חסרות עמודות בקובץ: {', '.join(heb[m] for m in missing)}. "
            f"נמצאו הכותרות: {', '.join(h for h in header if h)}. "
            f"בטמפלט צריך: מספר דרכון, שם מלא, מספר טלפון, סכום, מספר חשבון GMT")
    return mapping


def normalize_phone(raw):
    # תא מספרי מאקסל (openpyxl מחזיר float/int) — המרה מדויקת בלי לאבד ספרות
    if isinstance(raw, float) and raw.is_integer():
        raw = str(int(raw))
    elif isinstance(raw, int):
        raw = str(raw)
    s = str(raw or "").strip()
    # "5.01E+08" ב-CSV = אקסל דרס את המספר; הספרות האמיתיות אבדו — אין שחזור
    if re.fullmatch(r"\d+(\.\d+)?[eE][+\-]?\d+", s):
        return None, (f"המספר נשמר באקסל בפורמט מדעי ({raw}) והספרות אבדו — "
                      f"יש להגדיר את עמודת הטלפון כטקסט (Text) ולמלא מחדש")
    if s.endswith(".0"):
        s = s[:-2]
    s = re.sub(r"[\s\-().]", "", s)
    if s.startswith("+"):
        s = s[1:]
    if s.startswith("00972"):
        s = s[4:]
    if s.startswith("972"):
        digits = s[3:]
        if digits.startswith("0"):
            digits = digits[1:]
    elif s.startswith("0"):
        digits = s[1:]
    elif s.isdigit() and len(s) == 9 and s.startswith("5"):
        digits = s
    else:
        return None, f"מספר לא מזוהה: {raw}"
    if not (digits.isdigit() and len(digits) == 9 and digits.startswith("5")):
        return None, f"מספר לא תקין (צריך נייד ישראלי): {raw}"
    return "972" + digits, None


def parse_amount(raw):
    s = str(raw or "").replace("₪", "").replace(",", "").strip()
    try:
        val = round(float(s), 2)
    except ValueError:
        return None, f"סכום לא תקין: {raw}"
    if not (0 < val <= 5000):
        return None, f"סכום חריג ({val}) — בדיקה ידנית"
    return val, None


def _clean_id(raw) -> str:
    """דרכון / חשבון GMT: תא מספרי מאקסל → מחרוזת מדויקת; רווחים החוצה."""
    if isinstance(raw, float) and raw.is_integer():
        raw = str(int(raw))
    elif isinstance(raw, int):
        raw = str(raw)
    s = str(raw or "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    return re.sub(r"\s+", "", s)


def load_charges(blob: bytes, filename: str):
    """מחזיר (רשימת שורות תקינות, רשימת שגיאות)."""
    header, raw_rows = read_rows_from_bytes(blob, filename)
    cols = map_columns(header)
    good, bad, seen_phone, seen_pass = [], [], {}, {}
    for i, r in enumerate(raw_rows, start=2):
        def cell(f):
            idx = cols[f]
            return r[idx] if idx < len(r) else None
        passport = _clean_id(cell("passport"))
        name = str(cell("name") or "").strip()
        phone, perr = normalize_phone(cell("phone"))
        amount, aerr = parse_amount(cell("amount"))
        gmt = _clean_id(cell("gmt"))
        problems = [e for e in (perr, aerr) if e]
        if not name:
            problems.append("שם חסר")
        if not passport:
            problems.append("מספר דרכון חסר")
        elif re.fullmatch(r"\d+(\.\d+)?[eE][+\-]?\d+", str(cell("passport") or "").strip()):
            problems.append("מספר הדרכון נשמר בפורמט מדעי — להגדיר את העמודה כטקסט")
        if not gmt:
            problems.append("מספר חשבון GMT חסר")
        if passport and passport in seen_pass:
            problems.append(f"דרכון כפול בקובץ (שורה {seen_pass[passport]})")
        if phone and phone in seen_phone:
            problems.append(f"טלפון כפול בקובץ (שורה {seen_phone[phone]})")
        # רישום גם לשורות פסולות — כפילות מול שורה שגויה עדיין דורשת בדיקה אנושית
        if passport and passport not in seen_pass:
            seen_pass[passport] = i
        if phone and phone not in seen_phone:
            seen_phone[phone] = i
        if problems:
            bad.append({"line": i, "name": name, "passport": passport,
                        "raw_phone": str(cell("phone") or ""), "problems": problems})
        else:
            good.append({"line": i, "passport": passport, "name": name,
                         "phone": phone, "amount": amount, "gmt": gmt})
    return good, bad


# ------------------------------------------------------------- חשבונית ירוקה

class GiError(RuntimeError):
    pass


def gi_token() -> dict:
    key = os.getenv("SIM_BILLING_GI_KEY", "").strip()
    secret = os.getenv("SIM_BILLING_GI_SECRET", "").strip()
    if not key or not secret:
        raise GiError("חסרים מפתחות חשבונית ירוקה (SIM_BILLING_GI_KEY/SECRET)")
    r = requests.post(f"{GI_BASE}/account/token", json={"id": key, "secret": secret},
                      headers={"Content-Type": "application/json"}, timeout=30)
    if r.status_code >= 400:
        raise GiError(f"חשבונית ירוקה דחתה את המפתחות (HTTP {r.status_code})")
    jwt = (r.json() or {}).get("token") or r.headers.get("X-Authorization-Bearer", "")
    if not jwt:
        raise GiError("חשבונית ירוקה לא החזירה טוקן")
    return {"Authorization": f"Bearer {jwt}", "Content-Type": "application/json"}


def gi_verify_business(headers) -> str:
    """שער הבטיחות: החשבון חייב להיות ג.ד. פיקה, אחרת עצירה."""
    r = requests.get(f"{GI_BASE}/businesses/me", headers=headers, timeout=30)
    r.raise_for_status()
    biz = r.json() or {}
    tax_id = str(biz.get("taxId", ""))
    if tax_id != EXPECTED_TAX_ID:
        raise GiError(
            f"עצירה! המפתחות מחוברים ל-'{biz.get('name')}' (ח.פ {tax_id}) "
            f"ולא ל-{BUSINESS_NAME} (ח.פ {EXPECTED_TAX_ID}). אף מסמך לא הופק.")
    return biz.get("name", "")


def month_bounds_dates(month: str):
    y, m = (int(x) for x in month.split("-"))
    last = calendar.monthrange(y, m)[1]
    return f"{y}-{m:02d}-01", f"{y}-{m:02d}-{last:02d}"


def gi_billed(headers, month: str) -> tuple[set, set]:
    """(טלפונים, דרכונים) שכבר קיבלו מסמך החודש — מקור האמת למניעת כפל.
    הדרכון מחולץ מהערות המסמך ("דרכון: X") — fail-open אם לא נמצא."""
    frm, to = month_bounds_dates(month)
    phones, passports, page = set(), set(), 1
    while page <= 40:
        r = requests.post(f"{GI_BASE}/documents/search", headers=headers,
                          json={"fromDate": frm, "toDate": to, "type": [DOC_TYPE],
                                "pageSize": 100, "page": page}, timeout=30)
        r.raise_for_status()
        body = r.json() or {}
        items = body.get("items") or []
        for it in items:
            raw = ((it.get("client") or {}).get("phone")) or ""
            norm, _ = normalize_phone(raw)
            if norm:
                phones.add(norm)
            m = re.search(r"דרכון:\s*(\S+)", str(it.get("remarks") or ""))
            if m:
                passports.add(m.group(1))
        if len(items) < 100:
            break
        page += 1
    return phones, passports


def gi_billed_phones(headers, month: str) -> set:
    """תאימות לאחור — הטלפונים בלבד."""
    return gi_billed(headers, month)[0]


def split_vat(amount: float):
    if VAT_MODE == "exempt":
        return amount, 1, amount
    if VAT_MODE == "add":
        return amount, 0, round(amount * (1 + VAT_RATE), 2)
    base = round(amount / (1 + VAT_RATE), 2)
    for delta in (0, -0.01, 0.01, -0.02, 0.02):
        candidate = round(base + delta, 2)
        if round(candidate * (1 + VAT_RATE), 2) == amount:
            return candidate, 0, amount
    return base, 0, round(base * (1 + VAT_RATE), 2)


def _extract_url(body):
    url = body.get("url")
    if isinstance(url, dict):
        return url.get("origin") or url.get("he") or url.get("en") or ""
    if isinstance(url, str):
        return url
    return ""


def gi_create_doc(headers, row: dict, month: str):
    price, vat_type, total = split_vat(row["amount"])
    today = (datetime.now(IL_TZ) if IL_TZ else datetime.now()).date().isoformat()
    # הקבלה מכילה במפורש: דרכון, טלפון וחשבון GMT (בהערות המסמך המודפסות)
    remarks = (f"{DOC_REMARKS}\n"
               f"דרכון: {row.get('passport', '')} · טלפון: 0{row['phone'][3:]}"
               + (f" · חשבון GMT: {row.get('gmt', '')}" if row.get("gmt") else ""))
    payload = {
        "type": DOC_TYPE,
        "date": today,
        "lang": "he",
        "currency": "ILS",
        "remarks": remarks,
        "signed": True,
        "rounding": False,
        "client": {"name": row["name"], "phone": "0" + row["phone"][3:], "add": True},
        "income": [{"description": LINE_DESCRIPTION.format(month=month),
                    "quantity": 1, "price": price, "currency": "ILS",
                    "vatType": vat_type}],
        "payment": [{"date": today, "type": PAYMENT_TYPE, "price": total,
                     "currency": "ILS"}],
    }
    r = requests.post(f"{GI_BASE}/documents", headers=headers, json=payload, timeout=60)
    if r.status_code >= 400:
        raise GiError(f"חשבונית ירוקה דחתה את המסמך (HTTP {r.status_code}): {r.text[:200]}")
    body = r.json() or {}
    doc_id = body.get("id")
    url = _extract_url(body)
    if doc_id and not url:
        g = requests.get(f"{GI_BASE}/documents/{doc_id}", headers=headers, timeout=30)
        if g.ok:
            url = _extract_url(g.json() or {})
    if doc_id and not url:
        g = requests.get(f"{GI_BASE}/documents/{doc_id}/download/links", headers=headers, timeout=30)
        if g.ok:
            b2 = g.json() or {}
            url = b2.get("origin") or b2.get("url") or _extract_url(b2)
    if not doc_id:
        raise GiError(f"המסמך אולי נוצר אבל לא חזר מזהה: {str(body)[:150]}")
    return doc_id, body.get("number", ""), url or ""


# ------------------------------------------------------------------- שליחה

def send_message(phone: str, text: str) -> str:
    ch = channel()
    if ch == "whatsapp":
        inst = os.getenv("GREEN_API_BOT_INSTANCE_ID", "").strip()
        token = os.getenv("GREEN_API_BOT_TOKEN", "").strip()
        if not inst or not token:
            raise RuntimeError("חסרים GREEN_API_BOT_INSTANCE_ID/_TOKEN")
        # אינסטנסים חדשים של Green API מחייבים סאב-דומיין פר-אינסטנס (7107.api...)
        default_host = (f"https://{inst[:4]}.api.green-api.com"
                        if inst[:4].isdigit() else "https://api.green-api.com")
        host = os.getenv("GREEN_API_HOST", default_host).rstrip("/")
        r = requests.post(
            f"{host}/waInstance{inst}/sendMessage/{token}",
            json={"chatId": f"{phone}@c.us", "message": text}, timeout=30)
        r.raise_for_status()
        return "וואטסאפ"
    if ch == "inforu":
        user = os.getenv("INFORU_USER", "").strip()
        tok = os.getenv("INFORU_TOKEN", "").strip()
        if not user or not tok:
            raise RuntimeError("חסרים INFORU_USER/INFORU_TOKEN")
        r = requests.post(
            "https://capi.inforu.co.il/api/v2/SMS/SendSms",
            auth=(user, tok),
            json={"Data": {"Message": text,
                           "Recipients": [{"Phone": "0" + phone[3:]}],
                           "Settings": {"Sender": SMS_SENDER}}},
            timeout=30)
        r.raise_for_status()
        body = r.json() if r.content else {}
        status = body.get("StatusId", body.get("statusId"))
        if status not in (1, "1", None):
            raise RuntimeError(f"InforU החזיר סטטוס {status}")
        return "SMS"
    if ch == "dry":
        return "לא נשלחה הודעה (ערוץ ההודעות עדיין לא הופעל)"
    raise RuntimeError(f"ערוץ לא מוכר: {ch}")


# --------------------------------------------------------------------- ריצה

@dataclass
class RunState:
    run_id: str
    month: str
    total: int = 0
    done: int = 0
    status: str = "running"          # running | finished | failed
    error: str = ""
    started: str = ""
    results: list = field(default_factory=list)
    skipped: list = field(default_factory=list)

    @property
    def ok_count(self):
        return sum(1 for r in self.results if r["ok"])

    @property
    def fail_count(self):
        return sum(1 for r in self.results if not r["ok"])

    @property
    def ok_total(self):
        return sum(r["amount"] for r in self.results if r["ok"])


def execute_run(state: RunState, rows: list[dict], month: str, limit: int = 0):
    """מפיק מסמך ושולח הודעה לכל שורה. מעדכן את state תוך כדי (ל-polling)."""
    try:
        headers = gi_token()
        gi_verify_business(headers)
        billed_phones, billed_passports = gi_billed(headers, month)
        todo = []
        for r in rows:
            if r["phone"] in billed_phones or (r.get("passport") and r["passport"] in billed_passports):
                state.skipped.append({**r, "reason": "כבר קיבל מסמך החודש (לפי חשבונית ירוקה)"})
            else:
                todo.append(r)
        if limit:
            for r in todo[limit:]:
                state.skipped.append({**r, "reason": "מעבר למגבלת ריצת הניסיון"})
            todo = todo[:limit]
        state.total = len(todo)
        for r in todo:
            item = {"passport": r.get("passport", ""), "name": r["name"],
                    "phone": "0" + r["phone"][3:], "amount": r["amount"],
                    "gmt": r.get("gmt", ""),
                    "ok": False, "doc_number": "", "doc_url": "", "delivery": "", "error": ""}
            try:
                doc_id, doc_num, url = gi_create_doc(headers, r, month)
                item["doc_number"] = doc_num or doc_id
                item["doc_url"] = url
                msg = build_message(r["name"], f"{r['amount']:g}", month,
                                    url or "(link will follow)")
                try:
                    item["delivery"] = send_message(r["phone"], msg)
                except Exception as se:
                    item["delivery"] = f"המסמך הופק אך ההודעה נכשלה: {se}"
                item["ok"] = True
            except Exception as e:
                item["error"] = str(e)[:250]
            state.results.append(item)
            state.done += 1
        state.status = "finished"
    except Exception as e:
        state.error = str(e)[:400]
        state.status = "failed"
