"""
app.py — מערכת החיובים של ג.ד. פיקה הולדינגס בע"מ — ממשק ווב למנהל החיובים.

הזרימה: התחברות → הורדת טמפלט → העלאת קובץ חיובים → תצוגה מקדימה (ולידציה +
בדיקת כפל מול חשבונית ירוקה) → אישור אנושי מפורש → הפקת חשבוניות + הודעות →
דוח ריצה. כל ריצה מדווחת לטלגרם של דקל.

אבטחה: סיסמאות ב-env (ADMIN_PASSWORD לדקל, MANAGER_PASSWORD למנהל החיובים),
CSRF על כל POST, האטת כניסות כושלות (מפתוח לפי IP אמיתי מאחורי פרוקסי Render —
הלקח מתקרית הדשבורד 16/08), סשן חתום, אין סודות בקוד.
"""

from __future__ import annotations
import os
import io
import csv
import hmac
import time
import uuid
import threading
import urllib.request
import urllib.parse
from datetime import datetime

from flask import (Flask, render_template, request, redirect, url_for,
                   session, abort, Response, jsonify)

import engine

try:
    from zoneinfo import ZoneInfo
    IL_TZ = ZoneInfo("Asia/Jerusalem")
except Exception:
    IL_TZ = None

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "")
if not app.secret_key:
    # בלי מפתח סשן אין התחברות בטוחה — נופלים למפתח אקראי (סשנים יתאפסו ברענון תהליך)
    app.secret_key = uuid.uuid4().hex
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.getenv("COOKIE_SECURE", "1") == "1",
    MAX_CONTENT_LENGTH=5 * 1024 * 1024,
)

APP_NAME = "מערכת חיובים — ג.ד. פיקה"

# מצבים בזיכרון התהליך (Render free = תהליך יחיד). איבוד ב-restart = מעלים קובץ שוב.
PENDING: dict[str, dict] = {}     # token -> {rows, bad, filename, month, ts}
RUNS: dict[str, engine.RunState] = {}
_RUNS_LOCK = threading.Lock()
_FAILED_LOGINS: dict[str, list] = {}   # ip -> [timestamps]


# ------------------------------------------------------------------ עזרים

def now_il():
    return datetime.now(IL_TZ) if IL_TZ else datetime.now()


def client_ip() -> str:
    """ה-IP האמיתי מאחורי פרוקסי Render: הערך האחרון ב-XFF (חסין זיוף)."""
    xff = request.headers.get("X-Forwarded-For", "")
    if xff:
        return xff.split(",")[-1].strip()
    return request.remote_addr or "?"


def login_throttled(ip: str) -> bool:
    window = [t for t in _FAILED_LOGINS.get(ip, []) if time.time() - t < 900]
    _FAILED_LOGINS[ip] = window
    return len(window) >= 15


def record_failure(ip: str):
    _FAILED_LOGINS.setdefault(ip, []).append(time.time())


def csrf_token() -> str:
    if "_csrf" not in session:
        session["_csrf"] = uuid.uuid4().hex
    return session["_csrf"]


def check_csrf():
    tok = request.form.get("_csrf", "")
    if not tok or not hmac.compare_digest(tok, session.get("_csrf", "")):
        abort(400, "CSRF")


app.jinja_env.globals["csrf_token"] = csrf_token
app.jinja_env.globals["APP_NAME"] = APP_NAME


def logged_in() -> bool:
    return bool(session.get("role"))


def require_login():
    if not logged_in():
        return redirect(url_for("login"))
    return None


def notify_telegram(text: str):
    """דיווח לדקל על כל ריצה. נכשל בשקט — לא מפיל את המערכת."""
    token = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
    chat = os.getenv("TELEGRAM_CHAT_ID", "").strip()
    if not token or not chat:
        return
    try:
        data = urllib.parse.urlencode({"chat_id": chat, "text": text[:4000]}).encode()
        urllib.request.urlopen(
            urllib.request.Request(f"https://api.telegram.org/bot{token}/sendMessage",
                                   data=data), timeout=10)
    except Exception:
        pass


# ------------------------------------------------------------------ ראוטים

@app.get("/health")
def health():
    return {"ok": True, "app": "pika-billing", "channel": engine.channel(),
            "time": now_il().isoformat(timespec="seconds")}


@app.route("/login", methods=["GET", "POST"])
def login():
    error = ""
    if request.method == "POST":
        check_csrf()
        ip = client_ip()
        if login_throttled(ip):
            error = "יותר מדי ניסיונות. לנסות שוב בעוד רבע שעה."
        else:
            pw = request.form.get("password", "")
            admin = os.getenv("ADMIN_PASSWORD", "")
            manager = os.getenv("MANAGER_PASSWORD", "")
            role = None
            if admin and hmac.compare_digest(pw, admin):
                role = "admin"
            elif manager and hmac.compare_digest(pw, manager):
                role = "manager"
            if role:
                session.clear()
                session["role"] = role
                session["_csrf"] = uuid.uuid4().hex
                return redirect(url_for("dashboard"))
            record_failure(ip)
            time.sleep(1)
            error = "סיסמה שגויה"
    return render_template("login.html", error=error)


@app.get("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.get("/")
def dashboard():
    if (r := require_login()):
        return r
    month = now_il().strftime("%Y-%m")
    gi_ok, gi_msg, billed_count = False, "", 0
    try:
        headers = engine.gi_token()
        engine.gi_verify_business(headers)
        gi_ok = True
        billed_count = len(engine.gi_billed_phones(headers, month))
        gi_msg = "מחובר ומאומת: " + engine.BUSINESS_NAME
    except Exception as e:
        gi_msg = str(e)[:200]
    runs = sorted(RUNS.values(), key=lambda s: s.started, reverse=True)[:10]
    return render_template("dashboard.html", month=month, gi_ok=gi_ok, gi_msg=gi_msg,
                           billed_count=billed_count, channel=engine.channel(),
                           runs=runs, role=session.get("role"))


TEMPLATE_HEADERS = ["מספר דרכון", "שם מלא", "מספר טלפון", "סכום", "מספר חשבון GMT"]
TEMPLATE_ROWS = [
    ["N1234567", "Somchai Prasert", "0501234567", "55", "100200300"],
    ["EB0891011", "Ivan Petrov", "052-2345678", "60.5", "100200301"],
]


@app.get("/template.csv")
def template_csv():
    if (r := require_login()):
        return r
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(TEMPLATE_HEADERS)
    for row in TEMPLATE_ROWS:
        w.writerow(row)
    data = "﻿" + buf.getvalue()
    return Response(data, mimetype="text/csv",
                    headers={"Content-Disposition":
                             "attachment; filename=charges-template.csv"})


@app.get("/template.xlsx")
def template_xlsx():
    """טמפלט אקסל שבו עמודות הטלפון/דרכון/GMT כבר מוגדרות כטקסט —
    האפס המוביל שורד ואין 5.01E+08."""
    if (r := require_login()):
        return r
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "חיובים"
    ws.sheet_view.rightToLeft = True
    ws.append(TEMPLATE_HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="12263A")
    text_cols = (1, 3, 5)   # דרכון, טלפון, GMT — טקסט מפורש
    for row in TEMPLATE_ROWS:
        ws.append(row)
    for col_idx in text_cols:
        for row_idx in range(1, 201):    # גם שורות עתידיות שימולאו בקובץ
            ws.cell(row=row_idx, column=col_idx).number_format = "@"
    widths = [16, 24, 16, 10, 20]
    for i, wdt in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = wdt
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return Response(out.read(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition":
                             "attachment; filename=charges-template.xlsx"})


@app.post("/upload")
def upload():
    if (r := require_login()):
        return r
    check_csrf()
    f = request.files.get("charges")
    if not f or not f.filename:
        return render_template("dashboard_error.html", error="לא נבחר קובץ")
    month = request.form.get("month", "").strip() or now_il().strftime("%Y-%m")
    try:
        good, bad = engine.load_charges(f.read(), f.filename)
    except ValueError as e:
        return render_template("dashboard_error.html", error=str(e))
    already = []
    try:
        headers = engine.gi_token()
        engine.gi_verify_business(headers)
        billed = engine.gi_billed_phones(headers, month)
        already = [r_ for r_ in good if r_["phone"] in billed]
        good = [r_ for r_ in good if r_["phone"] not in billed]
    except Exception as e:
        return render_template("dashboard_error.html",
                               error=f"חשבונית ירוקה לא זמינה כרגע: {str(e)[:200]}")
    token = uuid.uuid4().hex
    PENDING[token] = {"rows": good, "bad": bad, "already": already,
                      "filename": f.filename, "month": month,
                      "has_gmt": any(r_.get("gmt") for r_ in good),
                      "ts": time.time()}
    # ניקוי העלאות ישנות מהזיכרון
    for k in [k for k, v in PENDING.items() if time.time() - v["ts"] > 3600]:
        PENDING.pop(k, None)
    return redirect(url_for("preview", token=token))


@app.get("/preview/<token>")
def preview(token):
    if (r := require_login()):
        return r
    p = PENDING.get(token)
    if not p:
        return render_template("dashboard_error.html",
                               error="ההעלאה פגה — להעלות את הקובץ שוב")
    total = sum(r_["amount"] for r_ in p["rows"])
    return render_template("preview.html", p=p, token=token, total=total,
                           channel=engine.channel())


@app.post("/run")
def start_run():
    if (r := require_login()):
        return r
    check_csrf()
    token = request.form.get("token", "")
    p = PENDING.pop(token, None)
    if not p:
        return render_template("dashboard_error.html",
                               error="ההעלאה פגה — להעלות את הקובץ שוב")
    if not p["rows"]:
        return render_template("dashboard_error.html", error="אין שורות תקינות להפקה")
    limit = 1 if request.form.get("trial") == "1" else 0
    run_id = uuid.uuid4().hex[:12]
    state = engine.RunState(run_id=run_id, month=p["month"],
                            started=now_il().isoformat(timespec="seconds"))
    with _RUNS_LOCK:
        RUNS[run_id] = state

    rows = p["rows"]

    def worker():
        engine.execute_run(state, rows, p["month"], limit=limit)
        mode = " (ריצת ניסיון — שורה אחת)" if limit else ""
        if state.status == "finished":
            notify_telegram(
                f"📱 חיובי סים {p['month']}{mode}\n"
                f"הופקו {state.ok_count} חשבוניות על סך {state.ok_total:,.2f} ₪ · "
                f"נכשלו {state.fail_count} · דולגו {len(state.skipped)}\n"
                f"קובץ: {p['filename']}")
        else:
            notify_telegram(f"❌ ריצת חיובי סים {p['month']} נעצרה: {state.error}")

    threading.Thread(target=worker, daemon=True).start()
    return redirect(url_for("run_page", run_id=run_id))


@app.get("/run/<run_id>")
def run_page(run_id):
    if (r := require_login()):
        return r
    state = RUNS.get(run_id)
    if not state:
        return render_template("dashboard_error.html", error="ריצה לא נמצאה")
    return render_template("run.html", state=state)


@app.get("/api/run/<run_id>")
def run_status(run_id):
    if (r := require_login()):
        return r
    state = RUNS.get(run_id)
    if not state:
        return jsonify({"error": "not found"}), 404
    return jsonify({"status": state.status, "done": state.done, "total": state.total,
                    "ok": state.ok_count, "fail": state.fail_count,
                    "error": state.error})


@app.route("/test-message", methods=["GET", "POST"])
def test_message():
    """אדמין בלבד: שליחת הודעת בדיקה אחת למספר שמוקלד — אימות ערוץ בלי חיובים."""
    if (r := require_login()):
        return r
    if session.get("role") != "admin":
        abort(403)
    result, error = "", ""
    if request.method == "POST":
        check_csrf()
        phone, perr = engine.normalize_phone(request.form.get("phone", ""))
        if perr:
            error = perr
        else:
            try:
                how = engine.send_message(
                    phone,
                    "הודעת בדיקה ממערכת החיובים של ג.ד. פיקה — הערוץ עובד ✓\n"
                    "Test message from G.D. Pika billing system — channel OK ✓")
                result = f"נשלח בהצלחה ({how}) אל 0{phone[3:]}"
            except Exception as e:
                error = f"השליחה נכשלה: {str(e)[:250]}"
    return render_template("test_message.html", result=result, error=error,
                           channel=engine.channel())


@app.get("/report/<run_id>")
def report(run_id):
    if (r := require_login()):
        return r
    state = RUNS.get(run_id)
    if not state:
        return render_template("dashboard_error.html",
                               error="הדוח לא בזיכרון (אחרי עדכון גרסה) — הנתונים המלאים בחשבונית ירוקה")
    return render_template("report.html", state=state)


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5480, debug=False)
